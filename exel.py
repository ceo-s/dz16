from openpyxl import Workbook



workbook = Workbook()
work_sheet = workbook.active
work_sheet.title = "data"

def create_xlsx(data: dict, file_name: str):
    path = f"static/results/{file_name}.xlsx"   

    for index, key in enumerate(data.keys()):
        work_sheet.cell(row=1, column=index+1, value=key)

    for index, val in enumerate(data.values()):
        for num, cel in enumerate(val):
            work_sheet.cell(row=num+2, column=index+1, value=cel)

    workbook.save(path)
    workbook.close()